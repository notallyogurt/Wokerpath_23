import json
import os
import pprint
import time
import openpyxl
import requests
from lxml.html import tostring
import pyperclip
# chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\selenium\\ChromeProfile"

def writeErrorData(url):
    with open('error.txt', 'a', encoding='utf-8-sig') as f:
        f.write(url + '\n')
        return
def re_return_str(listdata):
    if listdata:
        return listdata[0].replace('\n', ' ').replace('\t', '').replace(' ','').replace('\r','')
    else:
        return ''
# 删除excel当中重复数据 但需要每一个标题有内容
def remove_xlsx_double():
    wb = openpyxl.load_workbook('finly.xlsx')
    ws = wb.active
    finlydata = []
    for row in range(1, ws.max_row + 1):
        list_cell = []
        for cell in range(1, ws.max_column + 1):
            text = ws.cell(row, cell).value
            if not text:
                text = ''
            list_cell.append(text)
        finlydata.append(list_cell)
    wb.save('finly.xlsx')
    WB = openpyxl.Workbook()
    WS = WB.active
    hanlylist = []

    for rowlist in finlydata:
        savedata = []
        if rowlist[1] not in hanlylist and rowlist[1]:
            harldy = ''
            color = ''
            colot_list = []
            for chilrow in finlydata:
                if rowlist[1] == chilrow[1]:
                    savedata.append(chilrow)
                    harldy = chilrow[0]
                    color = chilrow[9]
                    colot_list.append(color)
                if not chilrow[1] and chilrow[0] == harldy and color == chilrow[9]:
                    savedata.append(chilrow)
                if chilrow[1] and chilrow[9]:
                    pass
            hanlylist.append(rowlist[1])
        for item in savedata:
            WS.append(item)
    WB.save("test.xlsx")
    # exit()

def return_secondTitleList(first_name, second_list):
    # first_name = 'Products'
    # second_list =['Conector', 'Baby Games', 'Science Games and Activities', 'Board Games', 'Educational Games', 'Magic', 'Arts & Crafts', 'Puzzles < 500 Pieces', 'Puzzles >= 500 Pieces', '3D Puzzles']
    return [first_name + f",{first_name}-{item}" for item in second_list]
def newxpath_list():
    str_1 = '''


/collections/road-gravel-bikes
/collections/mountain-bikes
/collections/kids-bikes
/collections/hybrid-bikes
/collections/urban-bikes

    '''
    list_1 = [item.rstrip().lstrip().capitalize() for item in str_1.split('\n') if item != '']
    print(len(list_1))
    print(list_1)


def xpathtolist():
    Base = ''
    # Base =  "-".join([item.capitalize() for item in "Products,Products-Board Games".split(',')])
    url = "https://www.decathlon.com"
    str_1 = """
/collections/surf-bodyboard
/collections/wetsuits-booties-hood
/collections/surf-accessories-spare-parts
/collections/beach-towels-ponchos
/collections/rashguards-uv-protection
"""

    str_1 = str_1.replace('\n', ',').replace('    ', '').rstrip().lstrip().replace(',,', ',').replace(',,', ',')

    if str_1[0] == ',':
        str_1 = str_1[1:]
    if str_1[-1] == ',':
        str_1 = str_1[:-1]
    list_1 = [item.rstrip().lstrip() for item in str_1.split(',')]
    list_1 = [item for item in list_1 if item != '']
    pprint.pprint(list_1)
    pyperclip.copy(str(list_1))
    if list_1[0].count('/') >= 2:
        list_1 = [url + item for item in list_1 if 'http' not in item]
        pyperclip.copy(str(list_1))
    if Base:
        list_1 = [Base +' '+ item for item in list_1]
    print(list_1)
    print(len(list_1))
    return list_1


def xpathgetHtml(data, res):
    code = res.apparent_encoding  # 获取url对应的编码格式
    try:
        text = tostring(data, encoding=code).decode(code)
    except Exception as e:
        print(data.xpath('//text()'), 'error data')
        text = ""
    return text
def getTagetitle(oncetitle,secondetitle,theardlist):
    oncetitle = oncetitle.replace('/','-').replace('  ',' ')
    if not theardlist:
        listdata = []
        for item in secondetitle:
            listdata.append(f"{oncetitle},{oncetitle}-{item}")
        print(listdata)
        return
    list_data = []
    if secondetitle:
        secondetitle = secondetitle.replace('/',' ').replace('-',' ').replace('_',' ').replace('  ',' ').replace('  ',' ')
        for item in theardlist:
            list_data.append(f"{oncetitle},{oncetitle}-{secondetitle},{oncetitle}-{secondetitle}-{item}")
    else:
        secondetitle = theardlist[0]
        secondetitle = secondetitle.replace('/', ' ').replace('-', ' ').replace('_', ' ').replace('  ',' ').replace('  ',' ')
        for item in theardlist[1:]:
            list_data.append(f"{oncetitle},{oncetitle}-{secondetitle},{oncetitle}-{secondetitle}-{item}")
    print(list_data)
    pyperclip.copy(str(list_data))
def delstr_json(data):
    if '.json' in data:
        return data.replace('.json','')
def resolve(filename):

    with open(f'{filename}.json', 'r', encoding='utf-8-sig') as f:
        jsdata = json.loads(f.read())

    finlysavelist = []
    saveidlist = []
    for item in jsdata:
        dict_finly = {'url': '', 'Collection': ''}
        for child in jsdata:
            if item['url'] == child['url']:
                dict_finly['url'] = item['url']
                for col in child['Collection'].split(','):
                    if col not in dict_finly['Collection']:
                        dict_finly['Collection'] += ',' + col
        if dict_finly['Collection'][0] == ',':
            dict_finly['Collection'] = dict_finly['Collection'][1:]
        # print(dict_finly)
        if dict_finly['url'] not in saveidlist:
            finlysavelist.append(dict_finly)
            saveidlist.append(dict_finly['url'])
    print("去重后长度为", len(saveidlist))
    with open(f"{filename}_finly.json", 'w', encoding='utf-8-sig') as f:
        f.write(json.dumps(finlysavelist, ensure_ascii=False))


def getxpathdata(listdata):
    if listdata == 0:
        return ""
    elif listdata == 1:
        return listdata[0].replace('  ', ' ').replace('\n', '').replace('\t', '').replace('€','').replace(',','.').rstrip().lstrip()
    else:
        return " ".join(listdata).replace('  ', ' ').replace('\n', '').replace('\t', '').replace('€','').replace(',','.').rstrip().lstrip()


def saveAllJson_fromUrl(filename, listdata):
    # filename = delstr_json(filename)
    with open(f'{filename}.json', 'a', encoding='utf-8-sig') as f:
        for item in listdata:
            dict_js = {}
            dict_js['url'] = item[0]
            dict_js['Collection'] = item[1]
            f.write(json.dumps(dict_js, ensure_ascii=False) + ',\n')


def loadjson(filename):
    filename = delstr_json(filename)
    with open(f'{filename}.json', 'r', encoding='utf-8-sig') as f:
        jsdata = json.loads(f.read())
        print("json 数据总长度", len(jsdata))
        return jsdata
def resovelNew(filename):


    jsdata = loadjson(filename)
    print("原始长度", len(jsdata))
    finlysavelist = []
    saveidlist = []
    for item in jsdata:
        for child in jsdata:
            if item['url'] == child['url']:
                for col in child['Collection'].split(','):
                    if col not in item['Collection']:
                        item['Collection'] += ',' + col
        if item['Collection'][0] == ',':
            item['Collection'] = item['Collection'][1:]
        if item['url'] not in saveidlist:
            finlysavelist.append(item)
            saveidlist.append(item['url'])
    print("去重后长度为", len(finlysavelist))
    filename = delstr_json(filename)
    with open(f"{filename}_finly.json", 'w', encoding='utf-8-sig') as f:
        f.write(json.dumps(finlysavelist, ensure_ascii=False))


def verdata_1():
    VariantSKU = ''
    # Variant Grams :固定为0
    VariantGrams = '0'
    # Variant Inventory Tracker ：固定为shopify
    VariantInventoryTracker = 'shopify'
    # Variant Inventory Qty：库存 固定为9999
    VariantInventoryQty = '9999'
    # # Variant Inventory Policy ：固定为deny
    VariantInventoryPolicy = 'deny'
    # # Variant Fulfillment Service :固定为manual
    VariantFulfillmentService = 'manual'
    # , VariantInventoryPolicy, VariantFulfillmentService
    return [VariantSKU, VariantGrams, VariantInventoryTracker, VariantInventoryQty, VariantInventoryPolicy,
            VariantFulfillmentService]


def verdata_2():
    # SEO Title,
    SEO_Title = ''
    # SEO Description,
    SEO_Description = ''
    # Google Shopping / Google Product Category,
    GoogleProductCategory = ''
    # Google Shopping / Gender,
    GoogleShoppingGender = ''
    # Google Shopping / Age,
    GoogleShoppingAge = ''
    # Group	Google Shopping / MPN,
    GoogleShopping_MPN = ''
    # Google Shopping / AdWords Grouping,
    Google_Shopping_AdWords_Groupin = ''
    # Google Shopping / AdWords Labels,
    GoogleShoppingAdWordsLabels = ''
    # Google Shopping / Condition,
    GoogleShoppingCondition = ''
    # Google Shopping / Custom Product,
    GoogleShoppingCustomProduct = ''
    # Google Shopping / Custom Label 0,
    GoogleShoppingCustomLabel0 = ''
    # Google Shopping / Custom Label 1,
    GoogleShoppingCustomLabel1 = ''
    # Google Shopping / Custom Label 2,
    GoogleShoppingCustomLabel2 = ''
    # Google Shopping / Custom Label 3,
    GoogleShoppingCustomLabel3 = ''
    # Google Shopping / Custom Label 4,
    GoogleShoppingCustomLabel4 = ''
    return [SEO_Title, SEO_Description, GoogleProductCategory, GoogleShoppingGender, GoogleShoppingAge,
            GoogleShopping_MPN, Google_Shopping_AdWords_Groupin, GoogleShoppingAdWordsLabels
        , GoogleShoppingCondition, GoogleShoppingCustomProduct, GoogleShoppingCustomLabel0, GoogleShoppingCustomLabel1,
            GoogleShoppingCustomLabel2, GoogleShoppingCustomLabel3, GoogleShoppingCustomLabel4]


def save_xlsx(filename, datalist):
    if not os.path.exists(f'{filename}.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        list_d = 'Handle,	Title,	Body (HTML),	Vendor,	Collection,	Type,	Tags,	Published,	Option1 Name,	Option1 Value,	Option2 Name,	Option2 Value,	Option3 Name,	Option3 Value,	Variant SKU,	Variant Grams,	Variant Inventory Tracker,	Variant Inventory Qty,	Variant Inventory Policy,	Variant Fulfillment Service,	Variant Price,	Variant Compare At Price,	Variant Requires Shipping,	Variant Taxable,	Variant Barcode,	Image Src,	Image Position,	Image Alt Text,	Gift Card,	SEO Title,	SEO Description,	Google Shopping / Google Product Category,	Google Shopping / Gender,	Google Shopping / Age, Group	Google Shopping / MPN,	Google Shopping / AdWords Grouping,	Google Shopping / AdWords Labels,	Google Shopping / Condition,	Google Shopping / Custom Product,	Google Shopping / Custom Label 0,	Google Shopping / Custom Label 1,	Google Shopping / Custom Label 2,	Google Shopping / Custom Label 3,	Google Shopping / Custom Label 4,	Variant Image,	Variant Weight Unit,	Variant Tax Code,	Cost per item,	url'.split(
            ',')
        ws.append(list_d)
        wb.save(f'{filename}.xlsx')
    wb = openpyxl.load_workbook(f'{filename}.xlsx')
    ws = wb.active
    # print(datalist[0][1:2])
    for itemlist in datalist:
        ws.append(itemlist)

    wb.save(f'{filename}.xlsx')


def return_img_infolist(imglist, infolist, foodprice_list, defult_list_1):
    return_list = []
    imgcount = 0
    for infochildlist in infolist:
        # 通过逗号进行拆分详情图和颜色关联信息
        # infochildlist[0][1] = infochildlist[0][1].split(',')[0]
        imginfo = infochildlist[0][1].split(',')[-1]
        # 通过颜色进行判断是否归属一张图片
        imgurl = [item for item in imglist if imginfo in item or imginfo.upper() in item ]
        # 如果图片数量小于选项数量
        if len(imgurl) <= len(infochildlist):
            for listchild in infochildlist:
                dataList = []
                try:
                    imgcount += 1
                    listchild[7] = imgcount
                    dataList += verdata_1() + foodprice_list + listchild + [
                        imgurl[infochildlist.index(listchild)]] + defult_list_1
                except:
                    imgcount -= 1
                    listchild[7] = ''
                    dataList += verdata_1() + foodprice_list + listchild + [''] + defult_list_1
                return_list.append(dataList)
        else:
            for lenindex in range(len(imgurl)):
                imgcount += 1
                dataList = []
                try:
                    infochildlist[lenindex][7] = imgcount
                    defult_list_1[0] = 'kg'
                    dataList += verdata_1() + foodprice_list + infochildlist[lenindex] + [
                        imgurl[lenindex]] + defult_list_1
                except:
                    dataList += ['' for i in range(6)] + ['', '', '', '', ''] + ['', '', '', '', '', '', '',
                                                                                 imgcount] + [
                                    imgurl[lenindex]] + defult_list_1
                return_list.append(dataList)
    finly_list = [item for item in return_list if item.count('') <= 18] + [item for item in return_list if
                                                                           item.count('') > 18]

    refinly_list = [item for item in finly_list]
    for index, item in enumerate(refinly_list):
        item[12] = item[12].split(',')[0]

    # print(return_list)
    return refinly_list


def getinfolist(option_1Name, option_1Value, option2_title, option2_value_title, option3_title, option3_value_title,
                optionbaseimg):
    info_list = []
    if len(option_1Value) > 0:
        imgpos = 0
        for option_1 in option_1Value:
            value1list = []
            if len(option2_value_title) > 0:
                for option_2 in option2_value_title:
                    if len(option3_value_title) > 0:
                        for option_3 in option3_value_title:
                            value1list.append([option_1Name, option_1, option2_title, option_2, option3_title, option_3,
                                               optionbaseimg[option_1Value.index(option_1)], str(imgpos)])
                    else:
                        value1list.append([option_1Name, option_1, option2_title, option_2, '', '',
                                           optionbaseimg[option_1Value.index(option_1)], str(imgpos)])
            else:
                value1list.append(
                    [option_1Name, option_1, '', '', '', '', optionbaseimg[option_1Value.index(option_1)], str(imgpos)])
            info_list.append(value1list)
    else:
        info_list.append(['', '', '', '', '', '', '', ''])

    return info_list


def request(url, headers=None, params=None, cookies=None, proxies=None, timeout=3):
    errcount = 0
    while True:
        try:
            res = requests.get(url=url, headers=headers, params=params, proxies=proxies, cookies=cookies, timeout=timeout)
            return res
        except Exception as e:
            print(e)
            if errcount > 3:
                print("程序异常次数过多，请检查网络或者联系程序员 13217252129")
                exit()
            errcount += 1
            time.sleep(2)


def baseInfo(filename, url, Body, Collection, Title, option_1Name, option_1Value, option2_title,
             option2_value_title, optionbaseimg, option_1imgList,
             VariantCompareAtPrice, VariantPrice, option3_title='', option3_value_title=[]):
    hanlder = url.split('/')[-1].replace(' ', '-').replace('.html', '')  # 商品所属组 相同hanlder 视为同一个商品
    if not VariantCompareAtPrice and VariantPrice:
        VariantCompareAtPrice = VariantPrice
    Collection = Collection.replace('_', ' ')
    Collection = ",".join([item.capitalize() for item in Collection.split(',')])
    if Collection:
        if Collection[0] == ",":
            Collection = Collection[1:]
    info_list = getinfolist(option_1Name, option_1Value, option2_title, option2_value_title, option3_title,
                            option3_value_title, optionbaseimg)

    # 商品售价  商品原价
    Variant_Requires_Shipping = "TRUE"
    Variant_Taxable = "FALSE"
    Variant_Barcode = ""
    foodprice_list = [VariantPrice, VariantCompareAtPrice, Variant_Requires_Shipping, Variant_Taxable, Variant_Barcode]

    # 商品默认字段
    VariantWeightUnit = 'kg'
    VariantTaxCode = ''
    Costperitem = ''
    defult_list_1 = [VariantWeightUnit, VariantTaxCode, Costperitem, url]
    img_infolist = return_img_infolist(option_1imgList, info_list, foodprice_list, defult_list_1)
    # print(img_infolist)
    Image_Alt_Text = ''
    Gift_Card = 'FALSE'
    Published = "TRUE"
    Vendor = ''
    Type = Collection.split(',')[0]
    Tags = Collection
    finlysavedata = []
    for item in img_infolist:
        if img_infolist.index(item) == 0:
            itemlist = [hanlder.lstrip().rstrip(), Title.lstrip().rstrip(), Body.lstrip().rstrip(), Vendor,
                        Collection.lstrip().rstrip(), Type.lstrip().rstrip(), Tags.lstrip().rstrip(), Published]
        else:
            itemlist = [hanlder, '', '', '', '', '', '', '']
        itemlist += item[11:-7] + item[:6] + item[6:11] + [item[-5], str(item[-6]), Image_Alt_Text] + [
            Gift_Card] + verdata_2() + [item[-7]] + item[-4:]
        itemlist[25] = itemlist[25].split('?')[0]
        itemlist[-5] = itemlist[-5].split('?')[0]
        finlysavedata.append(itemlist)
    save_xlsx(filename, finlysavedata)




if __name__ == '__main__':
    '''
    Collection = '3,4,5,5'  # 分类 用,进行分隔
    url = ''  # 产品链接
    Title = ''  # 标题
    Body = '' # 商品详情信息,需要对 button  超链接进行处理
    option_1Name = 'color'
    option_1Value = ['00000,333']  # 用逗号进行分割对应上每一张变量图片
    option2_title = 'size'
    option2_value_title = ['39', '40','41']
    optionbaseimg = ['imgbase',]    #商品主图列表   唯一一张 数量与颜色分类一致
    option_1imgList= ['333img1', ]  #产品变量图片列表  一个商品拥有多张图需注意 只有option_1Value 在图片存在才会找到
    VariantCompareAtPrice = ''    #原价
    VariantPrice = ''   #售价 若没有售价价格为  原价*.3


    baseInfo(filename, url, Body, Collection,  Title, option_1Name, option_1Value, option2_title, option2_value_title,optionbaseimg,option_1imgList,VariantCompareAtPrice,VariantPrice,option3_title='', option3_value_title=[])
    '''
    xpathtolist()
    # newxpath_list()